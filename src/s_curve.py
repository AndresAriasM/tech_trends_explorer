# src/s_curve.py
import streamlit as st
import os
import requests
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import time
from datetime import datetime
import re
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from utils.scopus_builder import scopus_equation_interface, parse_scopus_query
from scipy.optimize import curve_fit
import json
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Definir la función sigmoidal al nivel global para que esté disponible en toda la aplicación
def sigmoid(x, L, x0, k):
    """
    Función sigmoidal de tres parámetros.
    """
    return L / (1 + np.exp(-k * (x - x0)))

class ScopusPublicationsAnalyzer:
    """
    Clase para conectarse a la API de Scopus, buscar publicaciones académicas y analizarlas por año
    para crear una curva en S.
    """
    
    def __init__(self, api_key=None):
        """
        Inicializa el analizador.
        
        Args:
            api_key: La clave API de Scopus.
        """
        self.api_key = api_key
        self.base_url = "https://api.elsevier.com/content/search/scopus"
        self.headers = {
            "X-ELS-APIKey": api_key,
            "Accept": "application/json"
        }
        self.publications_by_year = {}
        
    def search_publications(self, query, max_results=1000):
        """
        Busca publicaciones en Scopus utilizando una consulta, sin filtros de año.
        
        Args:
            query: Consulta de búsqueda en formato de sintaxis de Scopus
            max_results: Número máximo de resultados a recuperar
            
        Returns:
            Resultados de la búsqueda o None si hay error
        """
        # Añadir filtro específico para excluir patentes
        query_papers = f"({query}) AND NOT (DOCTYPE(pt))"
        
        # Mostrar ecuación de búsqueda
        st.info(f"📝 Ecuación de búsqueda (Papers): {query_papers}")
        
        # Configurar parámetros de búsqueda
        base_url = "https://api.elsevier.com/content/search/scopus"
        headers = {
            "X-ELS-APIKey": self.api_key,
            "Accept": "application/json"
        }
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # Intentar la búsqueda
        try:
            # Verificar si la búsqueda funciona
            status_text.text("🔍 Verificando si la búsqueda funciona...")
            
            response = requests.get(
                base_url,
                headers=headers,
                params={"query": query_papers, "count": 1}
            )
            
            if response.status_code != 200:
                st.error(f"❌ Error en la búsqueda de papers: {response.status_code}")
                st.code(response.text)
                
                # Probar con una versión simplificada de la consulta para diagnosticar
                status_text.text("🔍 Probando con búsqueda simplificada...")
                simple_query = 'TITLE("banana") AND TITLE("flour") AND NOT (DOCTYPE(pt))'
                
                st.info(f"📝 Intentando con: {simple_query}")
                
                response = requests.get(
                    base_url,
                    headers=headers,
                    params={"query": simple_query, "count": 1}
                )
                
                if response.status_code == 200:
                    st.success("✅ La búsqueda simplificada funciona")
                    st.info("❗ El problema está en la complejidad de la ecuación original")
                    
                    # Usar la ecuación simplificada
                    query_papers = simple_query
                    st.info(f"📝 Nueva ecuación de búsqueda: {query_papers}")
                else:
                    st.error(f"❌ Error también con búsqueda simplificada: {response.status_code}")
                    st.code(response.text)
                    return None
            
            # Realizar la búsqueda con la ecuación (original o simplificada)
            status_text.text("🔍 Obteniendo resultados completos...")
            
            response = requests.get(
                base_url,
                headers=headers,
                params={"query": query_papers, "count": 25, "sort": "coverDate"}
            )
            
            if response.status_code == 200:
                data = response.json()
                total_results = int(data["search-results"]["opensearch:totalResults"])
                st.success(f"✅ Búsqueda exitosa: {total_results} papers encontrados")
                
                if total_results == 0:
                    status_text.text("❌ No se encontraron papers con esta ecuación")
                    return None
                
                # Recopilar todos los resultados
                all_results = []
                params = {"query": query_papers, "count": 25, "start": 0, "sort": "coverDate"}
                
                while params["start"] < min(total_results, max_results):
                    response = requests.get(base_url, headers=headers, params=params)
                    
                    if response.status_code == 200:
                        data = response.json()
                        entries = data["search-results"].get("entry", [])
                        all_results.extend(entries)
                        
                        # Actualizar progreso
                        progress = min(len(all_results) / min(total_results, max_results), 1.0)
                        progress_bar.progress(progress)
                        status_text.text(f"⏳ Recuperados {len(all_results)} de {min(total_results, max_results)} papers ({progress*100:.1f}%)...")
                        
                        params["start"] += params["count"]
                        
                        if params["start"] < min(total_results, max_results):
                            time.sleep(0.5)  # pequeña pausa entre solicitudes
                    else:
                        st.error(f"❌ Error al recuperar página: {response.status_code}")
                        st.code(response.text)
                        break
                
                # Limpieza y mensaje final
                progress_bar.empty()
                
                if all_results:
                    status_text.text(f"✅ Recuperados {len(all_results)} papers en total")
                    return all_results
                else:
                    status_text.text("❌ No se pudieron recuperar papers")
                    return None
            else:
                st.error(f"❌ Error en la búsqueda: {response.status_code}")
                st.code(response.text)
                return None
        except Exception as e:
            st.error(f"❌ Error inesperado en búsqueda de papers: {str(e)}")
            import traceback
            st.code(traceback.format_exc())
            return None
        
    def categorize_by_year(self, results):
        """
        Categoriza los resultados de publicaciones por año y completa los años faltantes con ceros.
        
        Args:
            results: Lista de resultados de la búsqueda
            
        Returns:
            Diccionario con el recuento por año
        """
        publications_by_year = {}
        
        st.text("📅 Categorizando publicaciones por año...")
        
        for entry in results:
            # Intentar extraer el año de diferentes campos
            year = None
            
            # Método 1: Del campo prism:coverDate
            if "prism:coverDate" in entry:
                try:
                    date_str = entry["prism:coverDate"]
                    year = datetime.strptime(date_str, "%Y-%m-%d").year
                except (ValueError, TypeError):
                    pass
            
            # Método 2: Del campo prism:publicationDate
            if not year and "prism:publicationDate" in entry:
                try:
                    date_str = entry["prism:publicationDate"]
                    year = datetime.strptime(date_str, "%Y-%m-%d").year
                except (ValueError, TypeError):
                    pass
                    
            # Método 3: Del texto de la fecha de portada
            if not year and "prism:coverDisplayDate" in entry:
                try:
                    year_str = entry["prism:coverDisplayDate"]
                    # Buscar un año en el texto con regex
                    match = re.search(r'(19|20)\d{2}', year_str)
                    if match:
                        year = int(match.group(0))
                except (ValueError, TypeError, AttributeError):
                    pass
            
            # Si no se pudo determinar el año, continuar con la siguiente entrada
            if not year:
                continue
            
            # Incrementar el contador para este año
            publications_by_year[year] = publications_by_year.get(year, 0) + 1
        
        # Completar años faltantes con ceros
        if publications_by_year:
            min_year = min(publications_by_year.keys())
            max_year = max(publications_by_year.keys())
            
            # Crear rango completo de años
            all_years = list(range(min_year, max_year + 1))
            
            # Completar con ceros los años faltantes
            complete_publications = {year: publications_by_year.get(year, 0) for year in all_years}
            
            # Ordenar por año
            self.publications_by_year = dict(sorted(complete_publications.items()))
        else:
            self.publications_by_year = {}
            
        return self.publications_by_year


class PatentDataManager:
    """
    Clase para gestionar datos de patentes cargados manualmente mediante archivos Excel.
    """
    
    @staticmethod
    def create_template(start_year=1920, end_year=None):
        """
        Crea una plantilla Excel para la carga de datos de patentes por año.
        
        Args:
            start_year: Año inicial para la plantilla
            end_year: Año final para la plantilla (por defecto es el año actual)
        
        Returns:
            BytesIO: Objeto de bytes con el contenido del archivo Excel
        """
        if end_year is None:
            end_year = datetime.now().year
        
        # Crear workbook y hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos de Patentes"
        
        # Añadir encabezados con estilo
        ws['A1'] = "Año"
        ws['B1'] = "Número de Patentes"
        ws['C1'] = "Notas/Observaciones"
        
        # Aplicar estilos a los encabezados
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}1']
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        
        # Añadir filas para cada año
        for i, year in enumerate(range(start_year, end_year + 1), start=2):
            ws[f'A{i}'] = year
            ws[f'B{i}'] = 0  # Valor predeterminado
            
            # Aplicar formato numérico a la columna de patentes
            ws[f'B{i}'].number_format = '0'
        
        # Añadir instrucciones en una nueva hoja
        ws_instructions = wb.create_sheet(title="Instrucciones")
        
        instructions = [
            "INSTRUCCIONES PARA COMPLETAR LA PLANTILLA DE PATENTES",
            "",
            "1. Completa la columna 'Número de Patentes' con la cantidad de patentes por año.",
            "2. Los datos deben ser valores numéricos enteros.",
            "3. No elimines ningún año de la lista.",
            "4. Puedes dejar años con valor 0 si no hay datos disponibles.",
            "5. La columna 'Notas/Observaciones' es opcional y puedes usarla para información adicional.",
            "6. Guarda el archivo cuando hayas completado los datos.",
            "7. Sube el archivo en la aplicación para analizar los datos de patentes.",
            "",
            "CÓMO OBTENER LOS DATOS DE PATENTES:",
            "",
            "- Opción 1: Busca en la pestaña 'Patents' de Scopus y anota el número de resultados por año.",
            "- Opción 2: Utiliza otras bases de datos de patentes como PatentScope, Google Patents, o Espacenet.",
            "- Opción 3: Consulta informes de oficinas de patentes nacionales o internacionales.",
            "",
            "Nota: Para una análisis preciso, asegúrate de utilizar la misma ecuación de búsqueda o criterios consistentes para todos los años."
        ]
        
        for i, line in enumerate(instructions, start=1):
            ws_instructions[f'A{i}'] = line
        
        # Ajustar ancho de columna en la hoja de instrucciones
        ws_instructions.column_dimensions['A'].width = 100
        
        # Guardar en un objeto de bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def load_data(uploaded_file):
        """
        Carga y procesa datos de patentes desde un archivo Excel subido.
        
        Args:
            uploaded_file: Archivo Excel subido mediante st.file_uploader
            
        Returns:
            dict: Diccionario con datos de patentes por año {año: conteo}
        """
        try:
            # Cargar el archivo Excel
            df = pd.read_excel(uploaded_file, sheet_name="Datos de Patentes")
            
            # Verificar que las columnas necesarias existen
            if "Año" not in df.columns or "Número de Patentes" not in df.columns:
                st.error("❌ El archivo no tiene el formato esperado. Asegúrate de usar la plantilla proporcionada.")
                return None
            
            # Convertir a diccionario {año: conteo}
            patents_by_year = dict(zip(df["Año"], df["Número de Patentes"]))
            
            # Filtrar años con valores nulos o no numéricos
            patents_by_year = {year: int(count) for year, count in patents_by_year.items() 
                              if pd.notna(count) and str(count).replace('.', '', 1).isdigit()}
            
            # Ordenar por año
            patents_by_year = dict(sorted(patents_by_year.items()))
            
            return patents_by_year
            
        except Exception as e:
            st.error(f"❌ Error al cargar el archivo: {str(e)}")
            return None


class TechnologyAnalyzer:
    """
    Clase para analizar datos de publicaciones o patentes y generar análisis de curva en S.
    """
    
    @staticmethod
    def analyze_s_curve(yearly_data):
        """
        Realiza un análisis completo de la curva en S con cálculos matemáticos.
        Versión mejorada con validación de datos suficientes.
        
        Args:
            yearly_data: Diccionario con conteo por año {año: conteo}
            
        Returns:
            DataFrame con los datos analizados, figura, info de ajuste y parámetros
        """
        if not yearly_data:
            st.error("No hay datos para analizar")
            return None, None, None, None
        
        # Verificar que hay suficientes datos
        if len(yearly_data) < 3:
            st.warning(f"⚠️ Solo hay {len(yearly_data)} años de datos. Se necesitan al menos 3 años para un análisis completo.")
            # Crear DataFrame básico sin análisis avanzado
            df = pd.DataFrame({
                'Año': list(yearly_data.keys()),
                'Cantidad': list(yearly_data.values())
            })
            df['Acumulado'] = df['Cantidad'].cumsum()
            
            # Crear gráfico simple
            fig = px.line(
                df, x='Año', y='Acumulado',
                title='Datos Acumulados por Año (Análisis Básico)',
                markers=True
            )
            fig.update_layout(
                plot_bgcolor='white',
                xaxis=dict(showgrid=True, gridcolor='lightgray'),
                yaxis=dict(showgrid=True, gridcolor='lightgray')
            )
            
            # Info básica
            basic_info = {
                'total_publications': sum(yearly_data.values()),
                'years_span': len(yearly_data),
                'avg_per_year': sum(yearly_data.values()) / len(yearly_data),
                'note': 'Datos insuficientes para análisis completo de curva S'
            }
            
            return df, fig, basic_info, None
        
        # Crear DataFrame base (solo si hay suficientes datos)
        df = pd.DataFrame({
            'Año': list(yearly_data.keys()),
            'Cantidad': list(yearly_data.values())
        })
        
        # Calcular acumulado
        df['Acumulado'] = df['Cantidad'].cumsum()
        
        # Establecer año como índice para cálculos
        df_analysis = df.copy()
        df_analysis.set_index('Año', inplace=True)
        
        # Calcular la tasa de crecimiento anual
        df_analysis['TasaCrecimiento'] = df_analysis['Acumulado'].pct_change() * 100
        
        # Calcular segunda derivada SOLO si hay suficientes datos
        if len(df_analysis) >= 3:
            try:
                df_analysis['SegundaDerivada'] = np.gradient(np.gradient(df_analysis['Acumulado']))
            except ValueError as e:
                st.warning(f"⚠️ No se pudo calcular la segunda derivada: {str(e)}")
                df_analysis['SegundaDerivada'] = 0
        else:
            df_analysis['SegundaDerivada'] = 0
        
        # Encontrar los años de puntos de inflexión exactos
        try:
            if len(df_analysis) >= 3 and 'SegundaDerivada' in df_analysis.columns:
                puntos_inflexion_exacto = df_analysis[df_analysis['SegundaDerivada'] == 0].index.tolist()
                if not puntos_inflexion_exacto and len(df_analysis) > 0:
                    # Si no hay puntos exactos con segunda derivada = 0, encontrar el más cercano
                    punto_cercano_idx = np.abs(df_analysis['SegundaDerivada']).idxmin()
                    puntos_inflexion_exacto = [punto_cercano_idx]
            else:
                puntos_inflexion_exacto = []
        except Exception as e:
            st.warning(f"⚠️ Error calculando puntos de inflexión: {str(e)}")
            puntos_inflexion_exacto = []
        
        # Preparar datos para el ajuste
        x_data = np.array(df_analysis.index)
        y_data = np.array(df_analysis['Acumulado'])
        
        # Ajustar la curva sigmoidal a los datos SOLO si hay suficientes puntos
        if len(x_data) >= 4:  # Mínimo 4 puntos para ajuste sigmoidal
            try:
                # Usar los mismos valores iniciales y parámetros 
                popt, pcov = curve_fit(
                    sigmoid, 
                    x_data, 
                    y_data, 
                    p0=[max(df_analysis['Acumulado']), np.median(df_analysis.index), 0.1], 
                    maxfev=5000
                )
                
                # Calcular los puntos de la curva ajustada
                curva_ajustada = sigmoid(x_data, *popt)
                df_analysis['Ajustada'] = curva_ajustada
                
                # Crear una tabla con datos relevantes
                df_parametros = pd.DataFrame({
                    'Parámetro': ['L', 'x0', 'k'],
                    'Valor ajustado': popt,
                    'Error estándar': np.sqrt(np.diag(pcov)),
                    'Valor T': popt/np.sqrt(np.diag(pcov)),
                    'Validación': ['Válido' if abs(valor_t) > 2 else 'No válido' 
                                for valor_t in popt/np.sqrt(np.diag(pcov))]
                })
                
                # Cálculo de R²
                r_squared = 1 - np.sum((y_data - curva_ajustada)**2) / np.sum((y_data - np.mean(y_data))**2)
                
                # Determinar la fase actual
                if len(df_analysis) >= 3:
                    ultima_derivada = df_analysis['SegundaDerivada'].iloc[-3:].mean() if len(df_analysis) >= 3 else 0
                else:
                    ultima_derivada = 0
                
                if ultima_derivada > 0:
                    fase = "Fase inicial (crecimiento acelerado)"
                    descripcion = "La tecnología está en su fase temprana con crecimiento acelerado."
                elif ultima_derivada < 0:
                    fase = "Fase de madurez (crecimiento desacelerado)"
                    descripcion = "La tecnología está madurando, el crecimiento se está desacelerando."
                else:
                    fase = "Punto de inflexión"
                    descripcion = "La tecnología está en el punto de inflexión entre el crecimiento acelerado y desacelerado."
                
                # Crear figura de Plotly con la curva acumulada y el ajuste
                fig = px.line(
                    df_analysis, x=df_analysis.index, y=['Acumulado', 'Ajustada'],
                    labels={'value': 'Cantidad Acumulada', 'variable': 'Tipo de Curva', 'Año': 'Año'},
                    title='Curva S - Acumulado por Año',
                    color_discrete_map={'Acumulado': 'blue', 'Ajustada': 'red'},
                    markers=True
                )
                
                # Añadir puntos de inflexión si existen
                for punto in puntos_inflexion_exacto:
                    fig.add_trace(
                        go.Scatter(
                            x=[punto],
                            y=[df_analysis.loc[punto, 'Acumulado']],
                            mode='markers',
                            name=f'Punto de inflexión ({punto})',
                            marker=dict(color='red', size=15, symbol='star')
                        )
                    )
                
                # Información de ajuste
                ajuste_info = {
                    'R2': r_squared,
                    'L': popt[0],
                    'x0': popt[1],
                    'k': popt[2],
                    'Fase': fase,
                    'Descripción': descripcion,
                    'Puntos_inflexion': puntos_inflexion_exacto
                }
                
            except Exception as e:
                st.warning(f"⚠️ Error en el ajuste de curva sigmoidal: {str(e)}")
                # Crear gráfico simple sin ajuste
                fig = px.line(
                    df_analysis, x=df_analysis.index, y='Acumulado',
                    title='Datos Acumulados por Año (Sin Ajuste)',
                    markers=True
                )
                
                # Info básica sin ajuste
                ajuste_info = {
                    'total_publications': sum(yearly_data.values()),
                    'years_span': len(yearly_data),
                    'avg_per_year': sum(yearly_data.values()) / len(yearly_data),
                    'note': 'No se pudo realizar ajuste sigmoidal'
                }
                df_parametros = None
        else:
            # No hay suficientes datos para ajuste sigmoidal
            st.warning(f"⚠️ Solo hay {len(x_data)} puntos de datos. Se necesitan al menos 4 para ajuste sigmoidal.")
            
            # Crear gráfico simple
            fig = px.line(
                df_analysis, x=df_analysis.index, y='Acumulado',
                title='Datos Acumulados por Año (Análisis Básico)',
                markers=True
            )
            
            # Info básica
            ajuste_info = {
                'total_publications': sum(yearly_data.values()),
                'years_span': len(yearly_data),
                'avg_per_year': sum(yearly_data.values()) / len(yearly_data),
                'note': 'Datos insuficientes para ajuste sigmoidal'
            }
            df_parametros = None
        
        # Mejorar aspecto visual del gráfico
        fig.update_layout(
            width=600,
            height=600,
            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
            plot_bgcolor='white',
            xaxis=dict(showgrid=True, gridcolor='lightgray'),
            yaxis=dict(showgrid=True, gridcolor='lightgray')
        )
        
        # Restaurar el índice para el DataFrame original
        df_analysis = df_analysis.reset_index()
        
        return df_analysis, fig, ajuste_info, df_parametros
    
    
    @staticmethod
    def display_data_table(yearly_data, title="Datos por Año"):
        """
        Muestra una tabla con los datos por año.
        
        Args:
            yearly_data: Diccionario {año: conteo}
            title: Título de la tabla
            
        Returns:
            DataFrame con los datos
        """
        if not yearly_data:
            st.error("No hay datos para mostrar")
            return None
        
        df = pd.DataFrame({
            'Año': list(yearly_data.keys()),
            'Cantidad': list(yearly_data.values())
        })
        
        # Calcular acumulado
        df['Acumulado'] = df['Cantidad'].cumsum()
        
        # Formatear años como enteros sin comas
        df['Año'] = df['Año'].astype(int)
        
        # Mostrar tabla
        st.write(f"### 📋 {title}")
        
        # Usar el componente nativo de Streamlit para tablas con estilo
        st.dataframe(
            df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Año": st.column_config.NumberColumn(format="%d"),
                "Cantidad": st.column_config.NumberColumn(format="%d"),
                "Acumulado": st.column_config.NumberColumn(format="%d")
            }
        )
        
        return df
    
    @staticmethod
    def display_s_curve_analysis(analysis_df, analysis_fig, ajuste_info, parametros, title="Análisis de Curva en S"):
        """
        Muestra el análisis completo de la curva en S.
        
        Args:
            analysis_df: DataFrame con los datos analizados
            analysis_fig: Figura de Plotly
            ajuste_info: Información del ajuste
            parametros: DataFrame con los parámetros del modelo
            title: Título del análisis
        """
        st.write(f"### 📊 {title}")
        
        # Mostrar gráfico de análisis
        if analysis_fig:
            st.plotly_chart(analysis_fig, use_container_width=True)
            
            # Mostrar información del ajuste
            if ajuste_info:
                st.write("#### Parámetros del Modelo Ajustado")
                
                # Crear columnas para mostrar métricas
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("R² del ajuste", f"{ajuste_info['R2']:.4f}")
                with col2:
                    st.metric("Punto de inflexión (x0)", f"{ajuste_info['x0']:.1f}")
                with col3:
                    st.metric("Máximo teórico (L)", f"{ajuste_info['L']:.1f}")
                
                # Mostrar parámetros del modelo
                if parametros is not None:
                    st.write("#### Parámetros del modelo sigmoidal")
                    
                    # Usar st.dataframe con estilo
                    st.dataframe(
                        parametros,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "Valor ajustado": st.column_config.NumberColumn("Valor ajustado", format="%.4f"),
                            "Error estándar": st.column_config.NumberColumn("Error estándar", format="%.4f"),
                            "Valor T": st.column_config.NumberColumn("Valor T", format="%.4f")
                        }
                    )
                
                # Mostrar fase actual
                st.info(f"**Fase actual de la tecnología**: {ajuste_info['Fase']}")
                st.write(ajuste_info['Descripción'])
    
    @staticmethod
    def export_data(df, filename_prefix, query):
        """
        Genera un botón para exportar los datos a CSV.
        
        Args:
            df: DataFrame con los datos
            filename_prefix: Prefijo para el nombre del archivo
            query: Consulta utilizada
        """
        if df is not None:
            st.write("### 📥 Exportar Datos")
            
            # Botón para descargar CSV
            csv = df.to_csv(index=False)
            query_term_clean = re.sub(r'[^\w\s]', '', query[:30])
            filename = f"{filename_prefix}_{query_term_clean}_{datetime.now().strftime('%Y%m%d')}.csv"
            
            st.download_button(
                label=f"📥 Descargar datos {filename_prefix} (CSV)",
                data=csv,
                file_name=filename,
                mime="text/csv",
                key=f"download_csv_{filename_prefix}_{int(time.time())}"  # ← KEY ÚNICO AGREGADO
            )

class OptimizedGooglePatentsAnalyzer:
    """
    Analizador optimizado de Google Patents que minimiza el consumo de tokens
    mediante estrategias inteligentes de búsqueda.
    """
    
    def __init__(self, api_key=None):
        self.api_key = api_key
        self.base_url = "https://serpapi.com/search"
        self.patents_by_year = {}
        self.cache = {}  # Cache para evitar búsquedas duplicadas
        
    def analyze_patents_optimized(self, query, start_year=None, end_year=None, max_tokens=50):
        """
        Análisis optimizado que minimiza el consumo de tokens usando múltiples estrategias.
        
        Args:
            query: Query de búsqueda original
            start_year: Año inicial (default: current_year - 25)
            end_year: Año final (default: current_year)
            max_tokens: Máximo número de requests permitidos
            
        Returns:
            dict: Patentes por año {año: cantidad}
        """
        current_year = datetime.now().year
        if not start_year:
            start_year = current_year - 25
        if not end_year:
            end_year = current_year
            
        st.info(f"🎯 **Estrategia Optimizada**: Máximo {max_tokens} requests para {end_year - start_year + 1} años")
        
        # PASO 1: Validar que la query funciona
        simplified_query = self._create_ultra_simple_query(query)
        if not self._validate_query_works(simplified_query):
            return None
            
        # PASO 2: Estrategia híbrida basada en rangos de años
        patents_by_year = {}
        tokens_used = 0
        
        # Definir estrategias por época
        strategies = [
            {
                'name': 'Época Antigua', 
                'range': (start_year, min(1999, end_year)),
                'strategy': 'bulk_search',  # Búsqueda masiva + procesamiento
                'batch_size': 20  # años por búsqueda
            },
            {
                'name': 'Época Moderna Temprana',
                'range': (max(2000, start_year), min(2009, end_year)),
                'strategy': 'decade_search',  # Por décadas
                'batch_size': 10
            },
            {
                'name': 'Época Reciente',
                'range': (max(2010, start_year), end_year),
                'strategy': 'smart_sampling',  # Muestreo inteligente
                'batch_size': 5
            }
        ]
        
        for strategy_info in strategies:
            if tokens_used >= max_tokens:
                st.warning(f"⚠️ Límite de tokens alcanzado ({max_tokens}). Deteniendo búsqueda.")
                break
                
            range_start, range_end = strategy_info['range']
            if range_start > range_end:
                continue
                
            st.write(f"### 📊 {strategy_info['name']} ({range_start}-{range_end})")
            
            strategy_result, tokens_consumed = self._execute_strategy(
                simplified_query, 
                range_start, 
                range_end, 
                strategy_info,
                max_tokens - tokens_used
            )
            
            if strategy_result:
                patents_by_year.update(strategy_result)
                tokens_used += tokens_consumed
                
            st.caption(f"✅ Tokens usados en esta época: {tokens_consumed} | Total: {tokens_used}/{max_tokens}")
        
        # PASO 3: Completar años faltantes con interpolación/extrapolación
        if patents_by_year:
            patents_by_year = self._fill_missing_years(patents_by_year, start_year, end_year)
            
        # PASO 4: Mostrar resumen de optimización
        self._show_optimization_summary(patents_by_year, tokens_used, max_tokens)
        
        return patents_by_year
    
    def _validate_query_works(self, query):
        """Valida que la query funciona con una búsqueda simple."""
        st.text("🔍 Validando query...")
        
        try:
            params = {
                "engine": "google_patents",
                "q": query,
                "api_key": self.api_key,
                "num": 10
            }
            
            response = requests.get(self.base_url, params=params, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                if "error" not in data:
                    total_results = data.get("search_information", {}).get("total_results", 0)
                    total_results = int(str(total_results).replace(",", "")) if total_results else 0
                    
                    if total_results > 0:
                        st.success(f"✅ Query válida: {total_results:,} resultados totales")
                        return True
                    else:
                        st.warning(f"⚠️ Query muy específica: 0 resultados. Usando query ultra-simple.")
                        return False
                else:
                    st.error(f"❌ Error en query: {data.get('error', 'Unknown')}")
                    return False
            else:
                st.error(f"❌ Error HTTP: {response.status_code}")
                return False
                
        except Exception as e:
            st.error(f"❌ Error validando query: {str(e)}")
            return False
    
    def _execute_strategy(self, query, start_year, end_year, strategy_info, max_tokens):
        """Ejecuta una estrategia específica de búsqueda."""
        strategy = strategy_info['strategy']
        
        if strategy == 'bulk_search':
            return self._bulk_search_strategy(query, start_year, end_year, max_tokens)
        elif strategy == 'decade_search':
            return self._decade_search_strategy(query, start_year, end_year, max_tokens)
        elif strategy == 'smart_sampling':
            return self._smart_sampling_strategy(query, start_year, end_year, max_tokens)
        else:
            return {}, 0
    
    def _bulk_search_strategy(self, query, start_year, end_year, max_tokens):
        """
        Estrategia de búsqueda masiva: Obtiene muchos resultados de una vez
        y extrae los años de los metadatos.
        """
        patents_by_year = {}
        tokens_used = 0
        
        if tokens_used >= max_tokens:
            return patents_by_year, tokens_used
        
        try:
            # Búsqueda amplia sin filtros de fecha
            params = {
                "engine": "google_patents",
                "q": f"{query} after:{start_year-1}-12-31 before:{end_year+1}-01-01",
                "api_key": self.api_key,
                "num": 100,  # Obtener más resultados por búsqueda
                "start": 0
            }
            
            st.text(f"🔄 Búsqueda masiva {start_year}-{end_year}...")
            
            # Hacer múltiples páginas si es necesario
            all_patents = []
            max_pages = min(5, max_tokens)  # Límite de páginas
            
            for page in range(max_pages):
                if tokens_used >= max_tokens:
                    break
                    
                params["start"] = page * 100
                response = requests.get(self.base_url, params=params, timeout=30)
                tokens_used += 1
                
                if response.status_code == 200:
                    data = response.json()
                    if "error" not in data and "organic_results" in data:
                        patents = data["organic_results"]
                        if not patents:  # No más resultados
                            break
                        all_patents.extend(patents)
                    else:
                        break
                else:
                    break
                    
                time.sleep(0.3)  # Rate limiting
            
            # Procesar resultados para extraer años
            if all_patents:
                for patent in all_patents:
                    year = self._extract_year_from_patent(patent)
                    if year and start_year <= year <= end_year:
                        patents_by_year[year] = patents_by_year.get(year, 0) + 1
                
                st.success(f"✅ Procesados {len(all_patents)} patentes, encontrados {sum(patents_by_year.values())} en rango")
            
        except Exception as e:
            st.warning(f"⚠️ Error en búsqueda masiva: {str(e)}")
        
        return patents_by_year, tokens_used
    
    def _decade_search_strategy(self, query, start_year, end_year, max_tokens):
        """Estrategia por décadas: Busca por rangos de 10 años."""
        patents_by_year = {}
        tokens_used = 0
        
        # Crear rangos de décadas
        decade_ranges = []
        current = start_year
        while current <= end_year:
            decade_end = min(current + 9, end_year)
            decade_ranges.append((current, decade_end))
            current += 10
        
        for decade_start, decade_end in decade_ranges:
            if tokens_used >= max_tokens:
                break
                
            try:
                params = {
                    "engine": "google_patents",
                    "q": f"{query} after:{decade_start-1}-12-31 before:{decade_end+1}-01-01",
                    "api_key": self.api_key,
                    "num": 50
                }
                
                st.text(f"🔄 Década {decade_start}-{decade_end}...")
                
                response = requests.get(self.base_url, params=params, timeout=30)
                tokens_used += 1
                
                if response.status_code == 200:
                    data = response.json()
                    if "error" not in data:
                        total_results = data.get("search_information", {}).get("total_results", 0)
                        total_results = int(str(total_results).replace(",", "")) if total_results else 0
                        
                        # Distribuir uniformemente los resultados en la década
                        if total_results > 0:
                            years_in_decade = decade_end - decade_start + 1
                            avg_per_year = total_results // years_in_decade
                            remainder = total_results % years_in_decade
                            
                            for i, year in enumerate(range(decade_start, decade_end + 1)):
                                patents_by_year[year] = avg_per_year + (1 if i < remainder else 0)
                
                time.sleep(0.3)
                
            except Exception as e:
                st.warning(f"⚠️ Error en década {decade_start}-{decade_end}: {str(e)}")
        
        return patents_by_year, tokens_used
    
    def _smart_sampling_strategy(self, query, start_year, end_year, max_tokens):
        """Estrategia de muestreo inteligente: Busca años clave y extrapola."""
        patents_by_year = {}
        tokens_used = 0
        
        # Seleccionar años clave para muestrear
        total_years = end_year - start_year + 1
        if total_years <= max_tokens:
            # Si hay suficientes tokens, buscar año por año
            sample_years = list(range(start_year, end_year + 1))
        else:
            # Muestrear años estratégicamente
            sample_size = min(max_tokens, total_years // 2)
            sample_years = self._select_strategic_years(start_year, end_year, sample_size)
        
        st.text(f"🎯 Muestreando {len(sample_years)} años clave de {total_years} totales")
        
        # Buscar años seleccionados
        for year in sample_years:
            if tokens_used >= max_tokens:
                break
                
            try:
                params = {
                    "engine": "google_patents",
                    "q": f"{query} after:{year-1}-12-31 before:{year+1}-01-01",
                    "api_key": self.api_key,
                    "num": 10
                }
                
                response = requests.get(self.base_url, params=params, timeout=30)
                tokens_used += 1
                
                if response.status_code == 200:
                    data = response.json()
                    if "error" not in data:
                        total_results = data.get("search_information", {}).get("total_results", 0)
                        total_results = int(str(total_results).replace(",", "")) if total_results else 0
                        patents_by_year[year] = total_results
                
                time.sleep(0.2)
                
            except Exception as e:
                st.warning(f"⚠️ Error en año {year}: {str(e)}")
        
        return patents_by_year, tokens_used
    
    def _select_strategic_years(self, start_year, end_year, sample_size):
        """Selecciona años estratégicos para muestrear."""
        years = list(range(start_year, end_year + 1))
        
        if sample_size >= len(years):
            return years
        
        # Estrategia: Incluir siempre primer año, último año, y distribuir el resto
        strategic_years = [start_year, end_year]
        
        if sample_size > 2:
            # Añadir años intermedios uniformemente distribuidos
            remaining_sample = sample_size - 2
            step = max(1, (end_year - start_year - 1) // (remaining_sample + 1))
            
            for i in range(1, remaining_sample + 1):
                year = start_year + i * step
                if year < end_year and year not in strategic_years:
                    strategic_years.append(year)
        
        return sorted(strategic_years)
    
    def _fill_missing_years(self, patents_by_year, start_year, end_year):
        """Completa años faltantes usando interpolación."""
        if not patents_by_year:
            return patents_by_year
        
        # Crear lista completa de años
        all_years = list(range(start_year, end_year + 1))
        filled_data = {}
        
        for year in all_years:
            if year in patents_by_year:
                filled_data[year] = patents_by_year[year]
            else:
                # Interpolación simple entre años conocidos
                filled_data[year] = self._interpolate_value(year, patents_by_year)
        
        return filled_data
    
    def _interpolate_value(self, target_year, known_data):
        """Interpola valor para un año faltante."""
        if not known_data:
            return 0
        
        known_years = sorted(known_data.keys())
        
        # Si el año objetivo está fuera del rango, usar extrapolación simple
        if target_year < min(known_years):
            return max(0, known_data[min(known_years)] // 2)  # Mitad del valor más temprano
        elif target_year > max(known_years):
            return known_data[max(known_years)]  # Mismo valor que el más reciente
        
        # Interpolación lineal entre dos puntos conocidos
        for i in range(len(known_years) - 1):
            year1, year2 = known_years[i], known_years[i + 1]
            if year1 <= target_year <= year2:
                value1, value2 = known_data[year1], known_data[year2]
                ratio = (target_year - year1) / (year2 - year1)
                return int(value1 + ratio * (value2 - value1))
        
        return 0
    
    def _extract_year_from_patent(self, patent_data):
        """Extrae el año de los metadatos de una patente."""
        # Intentar diferentes campos donde puede estar la fecha
        date_fields = [
            'publication_date', 
            'patent_date', 
            'filing_date',
            'priority_date'
        ]
        
        for field in date_fields:
            if field in patent_data:
                year = self._extract_year_from_date_string(patent_data[field])
                if year:
                    return year
        
        # Buscar en título o snippet
        text_to_search = f"{patent_data.get('title', '')} {patent_data.get('snippet', '')}"
        return self._extract_year_from_text(text_to_search)
    
    def _extract_year_from_date_string(self, date_str):
        """Extrae año de string de fecha."""
        if not date_str:
            return None
        
        # Buscar patrón de año (4 dígitos)
        import re
        match = re.search(r'\b(19|20)\d{2}\b', str(date_str))
        if match:
            year = int(match.group())
            current_year = datetime.now().year
            if 1900 <= year <= current_year:
                return year
        return None
    
    def _extract_year_from_text(self, text):
        """Extrae año de texto libre."""
        if not text:
            return None
        
        import re
        current_year = datetime.now().year
        years = re.findall(r'\b(19|20)\d{2}\b', text)
        
        if years:
            valid_years = [int(y) for y in years if 1900 <= int(y) <= current_year]
            if valid_years:
                return max(valid_years)  # Retornar el año más reciente
        return None
    
    def _create_ultra_simple_query(self, original_query):
        """Crea una query ultra-simple que tenga más probabilidad de funcionar."""
        # Extraer conceptos clave
        import re
        
        # Buscar términos entre comillas
        quoted_terms = re.findall(r'"([^"]+)"', original_query)
        
        if quoted_terms:
            # Usar solo el primer término y simplificar
            main_term = quoted_terms[0].strip()
            # Tomar solo la primera palabra del término principal
            words = main_term.split()
            if words:
                return f'"{words[0]}"'
        
        # Si no hay términos entre comillas, extraer palabras clave
        clean_query = re.sub(r'[()><=]', ' ', original_query)
        clean_query = re.sub(r'\b(AND|OR|NOT|TITLE|ABS|KEY)\b', ' ', clean_query)
        words = re.findall(r'\b[a-zA-Z]{4,}\b', clean_query)
        
        if words:
            return f'"{words[0]}"'
        
        # Fallback absoluto
        return "technology"
    
    def _show_optimization_summary(self, patents_by_year, tokens_used, max_tokens):
        """Muestra resumen de la optimización."""
        if not patents_by_year:
            st.error("❌ No se obtuvieron datos de patentes")
            return
        
        total_patents = sum(patents_by_year.values())
        years_covered = len(patents_by_year)
        
        st.success(f"🎯 **Optimización Exitosa**")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Tokens Usados", f"{tokens_used}/{max_tokens}")
        with col2:
            st.metric("Total Patentes", f"{total_patents:,}")
        with col3:
            st.metric("Años Cubiertos", years_covered)
        with col4:
            efficiency = total_patents / tokens_used if tokens_used > 0 else 0
            st.metric("Eficiencia", f"{efficiency:.1f} patentes/token")
        
        # Mostrar ahorro de tokens
        naive_tokens = years_covered  # Un token por año
        tokens_saved = naive_tokens - tokens_used
        if tokens_saved > 0:
            st.info(f"💰 **Ahorro**: {tokens_saved} tokens ({tokens_saved/naive_tokens*100:.1f}% menos que búsqueda año por año)")


# Función para integrar en el código existente
def integrate_optimized_patents_search():
    """
    Función para reemplazar la búsqueda de patentes existente.
    Usar esta función en lugar de GooglePatentsAnalyzer.search_patents()
    """
    
    def run_optimized_patents_analysis(scopus_query, start_year, end_year, serp_api_key, max_tokens=30):
        """
        Ejecuta análisis optimizado de patentes.
        
        Args:
            scopus_query: Query original de Scopus
            start_year: Año inicial
            end_year: Año final
            serp_api_key: API key de SerpAPI
            max_tokens: Máximo número de requests (default: 30)
            
        Returns:
            dict: Patentes por año
        """
        
        st.write("### 🚀 Análisis Optimizado de Patentes")
        
        # Configuración de optimización
        with st.expander("⚙️ Configuración de Optimización", expanded=True):
            col1, col2 = st.columns(2)
            
            with col1:
                max_tokens = st.slider(
                    "Máximo tokens a usar",
                    min_value=10,
                    max_value=100,
                    value=max_tokens,
                    help="Cada token = 1 request a SerpAPI"
                )
            
            with col2:
                years_span = end_year - start_year + 1
                naive_cost = years_span
                st.metric("Costo búsqueda naive", f"{naive_cost} tokens")
                st.metric("Costo optimizado", f"≤{max_tokens} tokens")
                
                if max_tokens < naive_cost:
                    savings = naive_cost - max_tokens
                    st.success(f"💰 Ahorro: {savings} tokens ({savings/naive_cost*100:.1f}%)")
        
        # Ejecutar análisis optimizado
        analyzer = OptimizedGooglePatentsAnalyzer(serp_api_key)
        patents_by_year = analyzer.analyze_patents_optimized(
            scopus_query, 
            start_year, 
            end_year, 
            max_tokens
        )
        
        return patents_by_year
    
    return run_optimized_patents_analysis    

def run_s_curve_analysis():
    """
    Ejecuta el análisis de curva en S para publicaciones académicas y patentes.
    """
    st.title("📈 Análisis de Curvas en S: Papers y Patentes")
    
    st.write("""
    Esta herramienta te permite analizar tendencias tecnológicas a lo largo del tiempo 
    utilizando datos de publicaciones académicas (papers) y patentes. El análisis de curvas en S 
    te ayuda a entender la fase de madurez en la que se encuentra una tecnología.
    """)
    
    # Configuración de APIs en el sidebar
    with st.sidebar:
        st.header("⚙️ Configuración")
        
        api_key_input = st.text_input(
            "API Key de Scopus",
            value=st.session_state.get('scopus_api_key', ''),
            type="password",
            help="Ingresa tu API key de Scopus/Elsevier"
        )
        
        use_default_key = st.checkbox(
            "Usar API key de ejemplo", 
            value=True, 
            help="Usa una API key de ejemplo para probar la aplicación"
        )
        
        if use_default_key:
            api_key = "113f57bcfb9e922c5a33ec02233ee24d"
        else:
            api_key = api_key_input
        
        st.session_state.scopus_api_key = api_key
        
        # Verificar SerpAPI para patentes
        serp_api_available = bool(st.session_state.get('serp_api_key'))
        if serp_api_available:
            st.success("✅ SerpAPI configurado para patentes")
        else:
            st.warning("⚠️ Configura SerpAPI para buscar patentes")
        
        # Opciones para tipos de datos
        st.subheader("Tipos de datos")
        analyze_papers = st.checkbox("Analizar papers", value=True)
        analyze_patents = st.checkbox("Analizar patentes", value=serp_api_available)
    
    # PESTAÑAS PRINCIPALES
    tab1, tab2 = st.tabs(["🔍 Análisis Automático", "📤 Carga Manual (Excel)"])
    
    # =================================================================
    # PESTAÑA 1: ANÁLISIS AUTOMÁTICO
    # =================================================================
    with tab1:
        st.header("🔍 Análisis Automático")
        st.write("Busca automáticamente papers en Scopus y patentes en Google Patents")
        
        # Pestañas para diferentes métodos de construcción de consulta
        query_tab1, query_tab2 = st.tabs(["Generador de Ecuaciones", "Ecuación Manual"])
        
        with query_tab1:
            # Utilizar el constructor de ecuaciones Scopus
            scopus_query = scopus_equation_interface()
        
        with query_tab2:
            # Entrada manual de ecuación
            manual_query = st.text_area(
                "Ecuación de búsqueda",
                placeholder='Ej: TITLE("Plantain" OR "banana" OR "musa") AND TITLE("flour" OR "starch")',
                height=100,
                key="manual_query_input_auto"
            )
            scopus_query = manual_query if manual_query else ""
        
        # Configuración de búsqueda
        with st.expander("⚙️ Configuración de Búsqueda", expanded=True):
            col1, col2 = st.columns(2)
            
            with col1:
                st.write("**Papers (Scopus):**")
                max_results = st.slider(
                    "Número máximo de resultados",
                    min_value=500,
                    max_value=10000,
                    value=2000,
                    step=100,
                    help="Mayor número = análisis más completo, pero toma más tiempo"
                )
            
            with col2:
                st.write("**Patentes (Google Patents):**")
                col2a, col2b = st.columns(2)
                
                with col2a:
                    patents_start_year = st.number_input(
                        "Año inicial", 
                        min_value=1950, 
                        max_value=datetime.now().year-1,
                        value=datetime.now().year - 20,
                        key="patents_start_year_main"
                    )
                    # Guardar en session_state
                    st.session_state['patents_start_year_config'] = patents_start_year
                
                with col2b:
                    patents_end_year = st.number_input(
                        "Año final", 
                        min_value=patents_start_year,
                        max_value=datetime.now().year,
                        value=datetime.now().year,
                        key="patents_end_year_main"
                    )
                    # Guardar en session_state
                    st.session_state['patents_end_year_config'] = patents_end_year
                
                years_span = patents_end_year - patents_start_year + 1
                st.caption(f"📡 {years_span} requests a SerpAPI (uno por año)")
        
        # BOTÓN PRINCIPAL DE ANÁLISIS
        if st.button("🔍 Analizar Tecnología", type="primary", use_container_width=True):
            if not scopus_query:
                st.error("Por favor, ingresa una ecuación de búsqueda válida")
                return
            
            if not analyze_papers and not analyze_patents:
                st.warning("Selecciona al menos un tipo de datos para analizar (papers o patentes).")
                return
            
            # Variables para almacenar resultados
            papers_by_year = None
            patents_by_year = None
            papers_analysis = None
            patents_analysis = None
            
            # =============================================================
            # 1. ANÁLISIS DE PAPERS
            # =============================================================
            if analyze_papers:
                st.write("## 📚 Análisis de Publicaciones Académicas (Papers)")
                
                if not api_key:
                    st.error("⚠️ Se requiere API Key de Scopus para analizar papers")
                else:
                    with st.spinner("🔄 Analizando publicaciones académicas..."):
                        try:
                            # Instanciar el analizador de papers
                            papers_analyzer = ScopusPublicationsAnalyzer(api_key)
                            
                            # Buscar papers
                            papers_results = papers_analyzer.search_publications(
                                scopus_query,
                                max_results=max_results
                            )
                            
                            # Verificar si se obtuvieron resultados
                            if papers_results:
                                # Categorizar por año
                                papers_by_year = papers_analyzer.categorize_by_year(papers_results)
                                
                                # Verificar si hay datos disponibles
                                if papers_by_year:
                                    # Mostrar tabla de papers por año
                                    df_papers = TechnologyAnalyzer.display_data_table(
                                        papers_by_year, 
                                        title="Tabla de Papers por Año"
                                    )
                                    
                                    # Realizar análisis de curva en S
                                    analysis_df, analysis_fig, ajuste_info, parametros = TechnologyAnalyzer.analyze_s_curve(papers_by_year)
                                    
                                    # Mostrar análisis de curva en S
                                    TechnologyAnalyzer.display_s_curve_analysis(
                                        analysis_df, 
                                        analysis_fig, 
                                        ajuste_info, 
                                        parametros,
                                        title="Análisis de Curva en S - Papers"
                                    )
                                    
                                    # Guardar para comparación posterior
                                    papers_analysis = {
                                        'df': analysis_df if analysis_df is not None else df_papers,
                                        'metrics': ajuste_info
                                    }
                                    
                                    # Exportar datos
                                    TechnologyAnalyzer.export_data(
                                        analysis_df if analysis_df is not None else df_papers,
                                        "papers",
                                        scopus_query
                                    )
                                else:
                                    st.warning("No se pudieron categorizar los papers por año.")
                            else:
                                st.error("No se pudieron obtener resultados de papers. Verifica tu API key y la ecuación de búsqueda.")
                        
                        except Exception as e:
                            st.error(f"Error en el análisis de papers: {str(e)}")
                            import traceback
                            st.code(traceback.format_exc())
            
            # =============================================================
            # 2. ANÁLISIS DE PATENTES - OPTIMIZADO
            # =============================================================
            if analyze_patents:
                st.write("## 📑 Análisis de Patentes (Optimizado)")
                
                if not st.session_state.get('serp_api_key'):
                    st.error("⚠️ Se requiere SerpAPI Key para analizar patentes. Configúrala en el panel lateral.")
                    patents_by_year = None
                    patents_analysis = None
                else:
                    # NUEVA CONFIGURACIÓN DE OPTIMIZACIÓN
                    with st.expander("🎯 Configuración de Optimización", expanded=True):
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            max_tokens = st.slider(
                                "Máximo tokens SerpAPI",
                                min_value=5,
                                max_value=100,
                                value=25,
                                help="Cada token = 1 request. Menos tokens = más optimización."
                            )
                        
                        with col2:
                            patents_start_year = st.session_state.get('patents_start_year_config', datetime.now().year - 15)
                            patents_end_year = st.session_state.get('patents_end_year_config', datetime.now().year)
                            years_span = patents_end_year - patents_start_year + 1
                            naive_cost = years_span
                            
                            st.metric("Años a analizar", years_span)
                            st.metric("Costo método naive", f"{naive_cost} tokens")
                        
                        with col3:
                            potential_savings = max(0, naive_cost - max_tokens)
                            savings_percent = (potential_savings / naive_cost * 100) if naive_cost > 0 else 0
                            
                            st.metric("Costo optimizado", f"≤{max_tokens} tokens")
                            if potential_savings > 0:
                                st.metric(
                                    "Ahorro estimado", 
                                    f"{potential_savings} tokens",
                                    delta=f"-{savings_percent:.1f}%"
                                )
                    
                    # EJECUTAR ANÁLISIS OPTIMIZADO
                    with st.spinner("🔄 Ejecutando análisis optimizado de patentes..."):
                        try:
                            # Usar la nueva clase optimizada
                            analyzer = OptimizedGooglePatentsAnalyzer(st.session_state.serp_api_key)
                            
                            patents_by_year = analyzer.analyze_patents_optimized(
                                scopus_query,
                                start_year=patents_start_year,
                                end_year=patents_end_year,
                                max_tokens=max_tokens
                            )
                            
                            # Verificar si se obtuvieron resultados
                            if patents_by_year and sum(patents_by_year.values()) > 0:
                                # Mostrar tabla de patentes por año
                                df_patents = TechnologyAnalyzer.display_data_table(
                                    patents_by_year, 
                                    title="Tabla de Patentes por Año (Optimizada)"
                                )
                                
                                # Realizar análisis de curva en S
                                analysis_df, analysis_fig, ajuste_info, parametros = TechnologyAnalyzer.analyze_s_curve(patents_by_year)
                                
                                # Mostrar análisis de curva en S
                                TechnologyAnalyzer.display_s_curve_analysis(
                                    analysis_df, 
                                    analysis_fig, 
                                    ajuste_info, 
                                    parametros,
                                    title="Análisis de Curva en S - Patentes (Optimizado)"
                                )
                                
                                # Guardar para comparación posterior
                                patents_analysis = {
                                    'df': analysis_df if analysis_df is not None else df_patents,
                                    'metrics': ajuste_info
                                }
                                
                                # Exportar datos
                                TechnologyAnalyzer.export_data(
                                    analysis_df if analysis_df is not None else df_patents,
                                    "patentes_optimizado",
                                    scopus_query
                                )
                                
                                # Mostrar estadísticas de optimización
                                st.success("✅ Análisis de patentes completado exitosamente")
                                
                            else:
                                st.warning("⚠️ No se obtuvieron resultados de patentes. Posibles causas:")
                                st.write("""
                                - La consulta es muy específica
                                - No existen patentes para esta tecnología en el rango de años
                                - Problema con la API de Google Patents
                                """)
                                
                                # Mostrar sugerencias para mejorar resultados
                                with st.expander("💡 Sugerencias para mejorar resultados"):
                                    st.write("""
                                    **Prueba estas estrategias:**
                                    
                                    1. **Simplifica la consulta**: Usa términos más generales
                                    2. **Amplía el rango de años**: Incluye años más antiguos
                                    3. **Incrementa el límite de tokens**: Permite más búsquedas
                                    4. **Verifica la tecnología**: Algunas tecnologías tienen pocas patentes
                                    """)
                                
                                patents_by_year = None
                                patents_analysis = None
                            
                        except Exception as e:
                            st.error(f"❌ Error en el análisis optimizado de patentes: {str(e)}")
                            
                            # Mostrar información de debug
                            with st.expander("🔧 Información de Debug"):
                                st.write("**Query original:**")
                                st.code(scopus_query)
                                
                                st.write("**Error completo:**")
                                import traceback
                                st.code(traceback.format_exc())
                            
                            patents_by_year = None
                            patents_analysis = None

            # 3. AÑADIR FUNCIÓN AUXILIAR PARA ANÁLISIS RÁPIDO DE VIABILIDAD
            def check_patents_viability(query, serp_api_key):
                """
                Función auxiliar para verificar rápidamente si vale la pena analizar patentes.
                Usa solo 1 token para verificar.
                """
                try:
                    analyzer = OptimizedGooglePatentsAnalyzer(serp_api_key)
                    simplified_query = analyzer._create_ultra_simple_query(query)
                    
                    # Hacer una búsqueda muy simple
                    params = {
                        "engine": "google_patents",
                        "q": simplified_query,
                        "api_key": serp_api_key,
                        "num": 1
                    }
                    
                    response = requests.get(analyzer.base_url, params=params, timeout=20)
                    
                    if response.status_code == 200:
                        data = response.json()
                        if "error" not in data:
                            total_results = data.get("search_information", {}).get("total_results", 0)
                            total_results = int(str(total_results).replace(",", "")) if total_results else 0
                            
                            return {
                                "viable": total_results > 0,
                                "total_results": total_results,
                                "simplified_query": simplified_query,
                                "recommendation": "Proceder con análisis" if total_results > 100 else 
                                            "Considerar ampliar términos" if total_results > 0 else 
                                            "Revisar consulta - muy específica"
                            }
                    
                    return {"viable": False, "error": "No se pudo conectar"}
                    
                except Exception as e:
                    return {"viable": False, "error": str(e)}

            # 4. INTEGRAR VERIFICACIÓN PREVIA EN LA INTERFAZ
            # Añadir esto ANTES del botón principal de análisis:

            # Verificación rápida de viabilidad
            if scopus_query and st.session_state.get('serp_api_key') and analyze_patents:
                with st.expander("🔍 Verificación Rápida de Patentes", expanded=False):
                    if st.button("🚀 Verificar viabilidad (1 token)", key="verify_patents"):
                        with st.spinner("Verificando..."):
                            viability = check_patents_viability(scopus_query, st.session_state.serp_api_key)
                            
                            if viability.get("viable"):
                                st.success(f"✅ {viability['recommendation']}")
                                st.info(f"📊 Aproximadamente {viability['total_results']:,} patentes encontradas")
                                st.caption(f"Query simplificada: {viability.get('simplified_query', 'N/A')}")
                            else:
                                st.warning(f"⚠️ {viability.get('recommendation', 'Problema detectado')}")
                                if "error" in viability:
                                    st.error(f"Error: {viability['error']}")

            # 5. CONFIGURACIÓN MEJORADA EN EL SIDEBAR
            # Añadir esto en la función sidebar_config():

            # En la sección de SerpAPI, añadir información sobre optimización:
            with st.expander("📈 SerpAPI (Hype Cycle & Patentes)", expanded=False):
                st.markdown('<div class="api-config">', unsafe_allow_html=True)
                serp_api_key = st.text_input(
                    "SerpAPI Key",
                    value=st.session_state.serp_api_key,
                    type="password",
                    help="Necesaria para análisis de Hype Cycle y búsqueda optimizada de patentes"
                )
                st.session_state.serp_api_key = serp_api_key
                
                # Información de optimización
                st.info("""
                **🎯 Optimización de Patentes:**
                - Búsqueda inteligente por rangos
                - Mínimo consumo de tokens
                - Interpolación de años faltantes
                - Estrategias adaptativas por época
                """)
                
                if st.button("🔄 Probar conexión SerpAPI", key="test_serp_unique"):
                    with st.spinner("Probando conexión con SerpAPI..."):
                        success, message = test_api_connection("serp", serp_api_key)
                        if success:
                            st.success(message)
                        else:
                            st.error(message)
                st.markdown('</div>', unsafe_allow_html=True)
                                
            # =============================================================
            # 3. COMPARACIÓN PAPERS vs PATENTES
            # =============================================================
            if papers_by_year and patents_by_year:
                st.write("## 🔄 Comparación Papers vs Patentes")
                
                with st.spinner("📊 Generando comparación..."):
                    # Crear DataFrame de comparación
                    compare_years = sorted(set(list(papers_by_year.keys()) + list(patents_by_year.keys())))
                    
                    compare_data = {
                        'Año': compare_years,
                        'Papers': [papers_by_year.get(year, 0) for year in compare_years],
                        'Patentes': [patents_by_year.get(year, 0) for year in compare_years]
                    }
                    
                    df_compare = pd.DataFrame(compare_data)
                    
                    # Calcular acumulados
                    df_compare['Papers Acumulados'] = df_compare['Papers'].cumsum()
                    df_compare['Patentes Acumuladas'] = df_compare['Patentes'].cumsum()
                    
                    # Mostrar tabla comparativa
                    st.write("### 📊 Tabla Comparativa")
                    st.dataframe(df_compare, use_container_width=True, hide_index=True)
                    
                    # Crear gráfico comparativo
                    fig_compare = go.Figure()
                    
                    # Añadir líneas para papers y patentes (acumulados)
                    fig_compare.add_trace(go.Scatter(
                        x=df_compare['Año'],
                        y=df_compare['Papers Acumulados'],
                        mode='lines+markers',
                        name='Papers (Acumulados)',
                        line=dict(color='blue', width=2)
                    ))
                    
                    fig_compare.add_trace(go.Scatter(
                        x=df_compare['Año'],
                        y=df_compare['Patentes Acumuladas'],
                        mode='lines+markers',
                        name='Patentes (Acumuladas)',
                        line=dict(color='red', width=2)
                    ))
                    
                    # Configurar aspecto visual
                    fig_compare.update_layout(
                        title="Comparación de Curvas S: Papers vs Patentes",
                        xaxis_title="Año",
                        yaxis_title="Cantidad Acumulada",
                        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
                        plot_bgcolor='white',
                        xaxis=dict(showgrid=True, gridcolor='lightgray'),
                        yaxis=dict(showgrid=True, gridcolor='lightgray')
                    )
                    
                    # Mostrar gráfico
                    st.plotly_chart(fig_compare, use_container_width=True)
                    
                    # =============================================================
                    # 4. ANÁLISIS DE BRECHA TEMPORAL (TIME LAG)
                    # =============================================================
                    st.write("### 🕰️ Análisis de Brecha Temporal (Time Lag)")
                    
                    if (papers_analysis and patents_analysis and 
                        papers_analysis['metrics'] and patents_analysis['metrics'] and
                        'x0' in papers_analysis['metrics'] and 'x0' in patents_analysis['metrics']):
                        
                        # Calcular brecha entre puntos de inflexión
                        papers_inflection = papers_analysis['metrics']['x0']
                        patents_inflection = patents_analysis['metrics']['x0']
                        time_lag = patents_inflection - papers_inflection
                        
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            st.metric(
                                "Punto de inflexión Papers", 
                                f"{papers_inflection:.1f}",
                                help="Año estimado del punto de inflexión para publicaciones académicas"
                            )
                        
                        with col2:
                            st.metric(
                                "Punto de inflexión Patentes", 
                                f"{patents_inflection:.1f}",
                                help="Año estimado del punto de inflexión para patentes"
                            )
                        
                        with col3:
                            delta_color = "normal" if abs(time_lag) < 1 else ("inverse" if time_lag > 0 else "off")
                            st.metric(
                                "Time Lag (años)", 
                                f"{time_lag:.1f}",
                                delta=f"{'+' if time_lag > 0 else ''}{time_lag:.1f}",
                                help="Diferencia temporal entre los puntos de inflexión"
                            )
                        
                        # Interpretación del time lag
                        if time_lag > 2:
                            st.info(f"🔍 **Interpretación**: Las patentes muestran un retraso significativo de {time_lag:.1f} años respecto a las publicaciones académicas. Esto sugiere que la investigación académica ha precedido considerablemente a la comercialización de esta tecnología.")
                        elif time_lag < -2:
                            st.info(f"🔍 **Interpretación**: Las patentes muestran un adelanto significativo de {abs(time_lag):.1f} años respecto a las publicaciones académicas. Esto sugiere que la comercialización ha precedido a la investigación académica profunda en esta tecnología.")
                        else:
                            st.info("🔍 **Interpretación**: No hay brecha temporal significativa entre publicaciones académicas y patentes, sugiriendo un desarrollo relativamente paralelo entre investigación y comercialización.")
                    else:
                        st.warning("No se pudieron calcular los puntos de inflexión para ambas curvas.")
                
                # Exportar comparación completa
                st.write("### 📥 Exportar Comparación Completa")
                combined_csv = df_compare.to_csv(index=False)
                st.download_button(
                    label="📊 Descargar Comparación Completa (CSV)",
                    data=combined_csv,
                    file_name=f"comparacion_papers_patents_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    mime="text/csv",
                    key="download_comparison_complete_csv"
                )
            
            # Opción para guardar análisis completo
            _show_save_analysis_option(scopus_query, papers_by_year, patents_by_year, papers_analysis, patents_analysis)
    
    # =================================================================
    # PESTAÑA 2: CARGA MANUAL (EXCEL)
    # =================================================================
    with tab2:
        st.header("📤 Carga Manual de Datos")
        st.write("""
        **Método auxiliar:** Si necesitas datos más específicos, de fuentes adicionales, 
        o quieres complementar los resultados automáticos, puedes usar plantillas Excel.
        """)
        
        # Subtabs para papers y patentes
        excel_tab1, excel_tab2 = st.tabs(["📄 Papers (Excel)", "📑 Patentes (Excel)"])
        
        with excel_tab1:
            st.write("### 📄 Carga Manual de Papers")
            st.info("Esta funcionalidad permite cargar datos de papers desde Excel si tienes fuentes adicionales.")
            st.write("*Próximamente: Plantilla para carga manual de papers*")
        
        with excel_tab2:
            st.write("### 📑 Carga Manual de Patentes")
            _show_excel_patents_interface()

# =================================================================
# FUNCIONES AUXILIARES
# =================================================================

def _show_save_analysis_option(query, papers_data, patents_data, papers_analysis, patents_analysis):
    """Muestra opciones para guardar el análisis completo"""
    if papers_data or patents_data:
        with st.expander("💾 Guardar Análisis Completo", expanded=False):
            st.write("Guarda este análisis para futuras comparaciones y referencias.")
            
            # Mostrar resumen de datos disponibles
            available_data = []
            if papers_data:
                available_data.append(f"✅ Papers: {sum(papers_data.values())} publicaciones")
            if patents_data:
                available_data.append(f"✅ Patentes: {sum(patents_data.values())} patentes")
            
            st.info("**Datos disponibles para guardar:**\n" + "\n".join(available_data))
            
            # Formulario para guardar
            with st.form("save_complete_analysis_form"):
                analysis_name = st.text_input(
                    "Nombre para este análisis",
                    value=f"Análisis completo - {query[:30]}..." if len(query) > 30 else f"Análisis completo - {query}"
                )
                
                save_method = st.radio(
                    "Método de guardado",
                    options=["Sistema de Base de Datos", "Archivo directo"],
                    index=0
                )
                
                submit = st.form_submit_button("💾 Guardar Análisis Completo", type="primary")
                
                if submit and analysis_name:
                    _save_complete_analysis(
                        analysis_name, query, papers_data, patents_data, 
                        papers_analysis, patents_analysis, save_method
                    )

def _show_excel_patents_interface():
    """Interfaz para carga manual de patentes via Excel"""
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("#### 📋 Descargar Plantilla")
        
        start_year = st.number_input(
            "Año inicial", 
            min_value=1900,
            max_value=datetime.now().year - 5,
            value=1920,
            key="excel_start_year"
        )
        
        end_year = st.number_input(
            "Año final", 
            min_value=start_year + 5,
            max_value=datetime.now().year,
            value=datetime.now().year,
            key="excel_end_year"
        )
        
        template_bytes = PatentDataManager.create_template(start_year, end_year)
        
        st.download_button(
            label="📥 Descargar Plantilla Excel",
            data=template_bytes,
            file_name=f"plantilla_patentes_{start_year}-{end_year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel_template_manual"
        )
    
    with col2:
        st.write("#### 📤 Subir Archivo")
        
        uploaded_file = st.file_uploader(
            "Selecciona archivo Excel",
            type=["xlsx", "xls"],
            key="excel_file_uploader_manual"
        )
        
        if uploaded_file is not None:
            patents_data = PatentDataManager.load_data(uploaded_file)
            
            if patents_data:
                st.success("✅ Archivo cargado correctamente")
                
                # Mostrar análisis inmediato
                df_patents = TechnologyAnalyzer.display_data_table(
                    patents_data, 
                    title="Datos de Patentes Cargados"
                )
                
                analysis_df, analysis_fig, ajuste_info, parametros = TechnologyAnalyzer.analyze_s_curve(patents_data)
                
                TechnologyAnalyzer.display_s_curve_analysis(
                    analysis_df, analysis_fig, ajuste_info, parametros,
                    title="Análisis de Curva en S - Patentes (Excel)"
                )
                
                TechnologyAnalyzer.export_data(
                    analysis_df if analysis_df is not None else df_patents,
                    "patentes_excel",
                    "datos_manuales"
                )

def _save_complete_analysis(name, query, papers_data, patents_data, papers_analysis, patents_analysis, method):
    """Guarda el análisis completo"""
    try:
        if method == "Sistema de Base de Datos":
            from data_storage import initialize_github_db
            
            db = initialize_github_db(use_local=True)
            if db:
                paper_metrics = papers_analysis['metrics'] if papers_analysis else None
                patent_metrics = patents_analysis['metrics'] if patents_analysis else None
                
                analysis_id = db.save_s_curve_analysis(
                    query=query,
                    paper_data=papers_data,
                    patent_data=patents_data,
                    paper_metrics=paper_metrics,
                    patent_metrics=patent_metrics,
                    analysis_name=name
                )
                
                if analysis_id:
                    st.success(f"✅ Análisis completo guardado con ID: {analysis_id}")
                else:
                    st.error("❌ Error al guardar en base de datos")
            else:
                st.error("❌ No se pudo inicializar el sistema de base de datos")
        
        else:  # Archivo directo
            from data_storage import save_analysis_direct
            
            paper_metrics = papers_analysis['metrics'] if papers_analysis else None
            patent_metrics = patents_analysis['metrics'] if patents_analysis else None
            
            result = save_analysis_direct(
                analysis_name=name,
                query=query,
                paper_data=papers_data,
                patent_data=patents_data,
                paper_metrics=paper_metrics,
                patent_metrics=patent_metrics
            )
            
            if result:
                st.success(f"✅ Análisis guardado en: {result}")
            else:
                st.error("❌ Error al guardar archivo directo")
                
    except Exception as e:
        st.error(f"❌ Error al guardar: {str(e)}")

def add_direct_save_option(analysis_name, query, paper_data=None, patent_data=None, paper_metrics=None, patent_metrics=None):
    """
    Añade una opción para guardar directamente los datos del análisis.
    """
    from data_storage import save_analysis_direct
    
    if st.button("💾 Guardar datos directamente", type="primary", key="save_direct_analysis"):
        result = save_analysis_direct(
            analysis_name=analysis_name,
            query=query,
            paper_data=paper_data,
            patent_data=patent_data,
            paper_metrics=paper_metrics,
            patent_metrics=patent_metrics
        )
        
        if result:
            st.success(f"Datos guardados correctamente en: {result}")
            # Mostrar opción para descargar el archivo
            try:
                with open(result, 'r', encoding='utf-8') as f:
                    file_content = f.read()
                    
                st.download_button(
                    label="📥 Descargar archivo JSON",
                    data=file_content,
                    file_name=os.path.basename(result),
                    mime="application/json",
                    key=f"download_json_direct_{int(time.time())}"  # ← KEY ÚNICO AGREGADO
                )
            except Exception as e:
                st.warning(f"No se pudo preparar descarga: {str(e)}")

if __name__ == "__main__":
    run_s_curve_analysis()